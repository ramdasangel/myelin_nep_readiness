#!/usr/bin/env python3
"""
Compute SRI (School Readiness Index) scores for all available constructs at branch level.

Constructs:
  C1 — Intent Readiness     (max 20)  from teacher+leader intent responses
  C2 — Practice Readiness   (max 25)  from teacher intent responses (depth analysis)
  C3 — Capacity Readiness   (max 20)  from MathTangle cognitive probe
  C4 — System Readiness     (max 20)  from baseline stage-2 survey
  C5 — Ecosystem Readiness  (max 15)  no data available → 0

Outputs:
  output/sri_branch_scores.csv
  Adds "SRI Scores" sheet to output/SRI_All_Constructs_Data.xlsx
"""

import csv
import json
import math
import os
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STAT = os.path.join(BASE, 'assets', 'myelin_stat_ro')
OUT  = os.path.join(BASE, 'output')

# ═══════════════════════════════════════════════════════════
#  Goal → Focus-Point (FP) mapping
# ═══════════════════════════════════════════════════════════

TEACHER_GOAL_FP = {
    # English
    'Each Child is Unique':          'FP1',
    'Competency-Based Learning':     'FP2',
    'Teacher Upskilling':            'FP3',
    'Diagnosing Learning Levels':    'FP4',
    'Parent–Teacher Collaboration':  'FP5',
    # Marathi (teacher & leader MR share these)
    'प्रत्येक मूल वेगळे आहे':       'FP1',
    'क्षमताधिष्ठित अध्ययन':         'FP2',
    'क्षमताधिष्टित अध्ययन':         'FP2',   # spelling variant
    'शिक्षक कौशल्यवृद्धी':          'FP3',
    'शिक्षक कौशल्यवृध्दी':          'FP3',   # spelling variant
    'अध्ययन स्तरांचे निदान':        'FP4',
    'पालक - शिक्षक सहयोग':          'FP5',
    'पालक -शिक्षक सहयोग':           'FP5',   # spacing variant
    'पालक - शिक्षक सहभागीता':       'FP5',   # alternate term
}

LEADER_EN_GOAL_FP = {
    # FP1 — Each Child is Unique / Understanding Learners
    'Learning Progress Monitoring':  'FP1',
    'Student Engagement':            'FP1',
    'Understanding Learners':        'FP1',
    # FP2 — Competency-Based Learning / Classroom Quality
    'Classroom Observation':         'FP2',
    'Activity Design':               'FP2',
    'Lesson Planning':               'FP2',
    'Task Rigor':                    'FP2',
    # FP3 — Teacher Upskilling / Professional Development
    'Teacher Support':               'FP3',
    'Teacher Coaching':              'FP3',
    'Innovation in Teaching':        'FP3',
    'Teacher Development':           'FP3',
    'Team Leadership':               'FP3',
    'School Improvement':            'FP3',
    # FP4 — Diagnosing Learning Levels / Data Use
    'Data Analysis':                 'FP4',
    'Instructional Review':          'FP4',
    'Learning Data Analysis':        'FP4',
    # FP5 — Parent-Teacher Collaboration
    'Parent Engagement':             'FP5',
    'Parent Communication':          'FP5',
    'School Communication':          'FP5',
}

# Leader Marathi uses same 5 goals as teacher Marathi
LEADER_MR_GOAL_FP = {k: v for k, v in TEACHER_GOAL_FP.items()}

FPS = ['FP1', 'FP2', 'FP3', 'FP4', 'FP5']

# ═══════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════

def read_csv(path):
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def cosine_sim(a, b):
    dot = sum(x * y for x, y in zip(a, b))
    ma = math.sqrt(sum(x * x for x in a))
    mb = math.sqrt(sum(x * x for x in b))
    return dot / (ma * mb) if ma > 0 and mb > 0 else 0.0


def safe_float(v, default=0.0):
    try:
        return float(v) if v and v.strip() else default
    except (ValueError, AttributeError):
        return default


def safe_int(v, default=0):
    try:
        return int(v) if v and v.strip() else default
    except (ValueError, AttributeError):
        return default


def build_branch_name_to_code():
    """Build branchName → branchCode from all available response CSVs."""
    mapping = {}
    sources = [
        os.path.join(STAT, 'system_readiness_responses_detail.csv'),
        os.path.join(STAT, 'intent_teacher_english_responses.csv'),
        os.path.join(STAT, 'intent_teacher_marathi_responses.csv'),
        os.path.join(STAT, 'intent_leader_english_responses.csv'),
        os.path.join(STAT, 'intent_leader_marathi_responses.csv'),
    ]
    for src in sources:
        if not os.path.exists(src):
            continue
        for row in read_csv(src):
            bc = row.get('branchCode', '').strip()
            bn = row.get('branchName', '').strip()
            if bc and bn and bn not in mapping:
                mapping[bn] = bc
    return mapping


def build_code_to_name():
    """Build branchCode → (branchName, schoolName) lookup."""
    mapping = {}
    src = os.path.join(STAT, 'system_readiness_responses_detail.csv')
    if os.path.exists(src):
        for row in read_csv(src):
            bc = row['branchCode'].strip()
            bn = row['branchName'].strip()
            sn = row.get('schoolName', '').strip()
            if bc and bc not in mapping:
                mapping[bc] = (bn, sn)
    # Also from intent (fills gaps for branches not in system data)
    for fn in ['intent_teacher_english_responses.csv',
               'intent_teacher_marathi_responses.csv',
               'intent_leader_english_responses.csv',
               'intent_leader_marathi_responses.csv']:
        p = os.path.join(STAT, fn)
        if not os.path.exists(p):
            continue
        for row in read_csv(p):
            bc = row['branchCode'].strip()
            bn = row['branchName'].strip()
            sn = row.get('schoolName', '').strip()
            if bc and bc not in mapping:
                mapping[bc] = (bn, sn)
    return mapping


# ═══════════════════════════════════════════════════════════
#  C1 — Intent Readiness (max 20)
# ═══════════════════════════════════════════════════════════

def compute_c1():
    """
    ADV (Aspirational Depth Vector): for each FP, the mean selectedOption / 3
    gives a [0,1] depth score.

    Coverage  = fraction of FPs with ADV > τ  (τ = 0.10)
    Balance   = 1 − Var(ADV) / 0.16           (0.16 = max variance for 5 bins)
    Alignment = cosine similarity between teacher ADV and leader ADV

    TeacherScore = TeacherCoverage × TeacherBalance × Alignment
    LeaderScore  = LeaderCoverage  × LeaderBalance  × Alignment
    C1 = 10 × TeacherScore  +  10 × LeaderScore
    """
    # Accumulate: branch → role → fp → [selectedOption values]
    teacher_fp = defaultdict(lambda: defaultdict(list))  # [bc][fp] → [opts]
    leader_fp  = defaultdict(lambda: defaultdict(list))

    intent_files = [
        ('intent_teacher_english_responses.csv',  'Teacher', TEACHER_GOAL_FP),
        ('intent_teacher_marathi_responses.csv',   'Teacher', TEACHER_GOAL_FP),
        ('intent_leader_english_responses.csv',    'Leader',  LEADER_EN_GOAL_FP),
        ('intent_leader_marathi_responses.csv',    'Leader',  LEADER_MR_GOAL_FP),
    ]

    unmapped_goals = set()
    for fn, role, goal_map in intent_files:
        path = os.path.join(STAT, fn)
        if not os.path.exists(path):
            print(f"  SKIP (not found): {fn}")
            continue
        for row in read_csv(path):
            bc   = row['branchCode'].strip()
            goal = row['goal'].strip()
            opt  = safe_int(row.get('selectedOption', ''), -1)
            if opt < 0 or not bc:
                continue
            fp = goal_map.get(goal)
            if not fp:
                unmapped_goals.add(goal)
                continue
            if role == 'Teacher':
                teacher_fp[bc][fp].append(opt)
            else:
                leader_fp[bc][fp].append(opt)

    if unmapped_goals:
        print(f"  Warning: {len(unmapped_goals)} unmapped goal(s): "
              f"{list(unmapped_goals)[:3]}...")

    all_branches = sorted(set(teacher_fp) | set(leader_fp))
    scores = {}

    for bc in all_branches:
        # Teacher ADV
        t_adv = []
        for fp in FPS:
            vals = teacher_fp[bc].get(fp, [])
            t_adv.append(sum(vals) / (len(vals) * 3.0) if vals else 0.0)

        # Leader ADV
        l_adv = []
        for fp in FPS:
            vals = leader_fp[bc].get(fp, [])
            l_adv.append(sum(vals) / (len(vals) * 3.0) if vals else 0.0)

        # Coverage (fraction of FPs above threshold)
        tau = 0.10
        t_cov = sum(1 for v in t_adv if v > tau) / 5.0
        l_cov = sum(1 for v in l_adv if v > tau) / 5.0

        # Balance (1 − variance / max_variance)
        def variance(vec):
            if not vec or all(v == 0 for v in vec):
                return 0.16  # treat all-zero as max variance
            mu = sum(vec) / len(vec)
            return sum((v - mu) ** 2 for v in vec) / len(vec)

        t_bal = max(0.0, 1 - variance(t_adv) / 0.16)
        l_bal = max(0.0, 1 - variance(l_adv) / 0.16)

        # Alignment
        align = cosine_sim(t_adv, l_adv)

        # Scores
        t_score = t_cov * t_bal * align
        l_score = l_cov * l_bal * align
        c1 = 10 * t_score + 10 * l_score

        scores[bc] = {
            'C1_TeacherScore': round(t_score, 4),
            'C1_LeaderScore':  round(l_score, 4),
            'C1_Alignment':    round(align, 4),
            'C1':              round(c1, 2),
        }

    return scores


# ═══════════════════════════════════════════════════════════
#  C2 — Practice Readiness (max 25)
# ═══════════════════════════════════════════════════════════

def compute_c2():
    """
    Uses TEACHER intent responses only.
    selectedOption 0-3 → depth D1-D4.

    PDDM (Practice Depth Distribution Matrix): 5 × 4 matrix
    Non-observable cells (masked to 0):
      FP1-D1 (opt 0), FP1-D4 (opt 3), FP2-D4 (opt 3)

    Expected depth per FP = weighted average of depth levels.
    Normalize E to achievable range per FP.
    DepthEffectiveness = CoverageFP × mean(E_norm)
    C2 = 25 × DepthEffectiveness
    """
    NON_OBS = {('FP1', 0), ('FP1', 3), ('FP2', 3)}

    # Achievable depth ranges after masking
    # FP1: only D2 (opt1) and D3 (opt2) remain → max weighted avg = 2 (all opt2)
    # FP2: D1-D3 remain (opt 0,1,2) → max weighted avg = 2 (all opt2)
    # FP3-FP5: all D1-D4 (opt 0-3) → max weighted avg = 3
    MAX_DEPTH = {'FP1': 2.0, 'FP2': 2.0, 'FP3': 3.0, 'FP4': 3.0, 'FP5': 3.0}

    # Accumulate: branch → fp → depth → count
    branch_fp_depth = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for fn in ['intent_teacher_english_responses.csv',
               'intent_teacher_marathi_responses.csv']:
        path = os.path.join(STAT, fn)
        if not os.path.exists(path):
            continue
        for row in read_csv(path):
            bc   = row['branchCode'].strip()
            goal = row['goal'].strip()
            opt  = safe_int(row.get('selectedOption', ''), -1)
            if opt < 0 or not bc:
                continue
            fp = TEACHER_GOAL_FP.get(goal)
            if not fp:
                continue
            branch_fp_depth[bc][fp][opt] += 1

    scores = {}
    for bc in branch_fp_depth:
        e_norm_values = []
        covered = 0

        for fp in FPS:
            raw = [branch_fp_depth[bc][fp].get(d, 0) for d in range(4)]
            # Apply observability mask
            masked = [0 if (fp, d) in NON_OBS else raw[d] for d in range(4)]
            total = sum(masked)
            if total == 0:
                continue
            covered += 1
            # Expected depth = weighted average of option indices
            e_depth = sum(d * masked[d] for d in range(4)) / total
            e_n = min(e_depth / MAX_DEPTH[fp], 1.0)
            e_norm_values.append(e_n)

        cov_fp = covered / 5.0
        mean_e = sum(e_norm_values) / len(e_norm_values) if e_norm_values else 0.0
        depth_eff = cov_fp * mean_e
        c2 = 25 * depth_eff

        scores[bc] = {
            'C2_DepthEffectiveness': round(depth_eff, 4),
            'C2': round(c2, 2),
        }

    return scores


# ═══════════════════════════════════════════════════════════
#  C3 — Capacity Readiness (max 20)
# ═══════════════════════════════════════════════════════════

def compute_c3():
    """
    From MathTangle master data, compute 4 CPD dimensions per user:
      CPD1 (Higher-order)  = (Analyze_R + Evaluate_R) / all_cog_total
      CPD2 (Decomposition) = (L2_acc + L3_acc) / 2
      CPD3 (Revision)      = 1 − IndecisionScore / max_indecision
      CPD4 (Consistency)   = 1 − FlipFlops / TotalQuestions
    CPI = mean(CPD1..CPD4)
    C3 = 20 × CPI_branch (mean CPI across users in branch)
    """
    branch_map = build_branch_name_to_code()
    path = os.path.join(OUT, 'mathangle_master.csv')
    if not os.path.exists(path):
        print("  SKIP: mathangle_master.csv not found")
        return {}

    rows = read_csv(path)

    # First pass: collect per-user data and find max indecision
    user_records = []
    max_indecision = 0.0

    for row in rows:
        branch = row['Branch'].strip()
        bc = branch_map.get(branch, '')

        # Cognitive right counts
        analyze_r  = safe_int(row.get('Cog:Analyze (Right)', ''))
        evaluate_r = safe_int(row.get('Cog:Evaluate (Right)', ''))

        # Total cognitive questions
        cog_total = (safe_int(row.get('Cog:Understand (Total)', ''))
                   + safe_int(row.get('Cog:Apply (Total)', ''))
                   + safe_int(row.get('Cog:Analyze (Total)', ''))
                   + safe_int(row.get('Cog:Evaluate (Total)', '')))

        # CPD1
        cpd1 = (analyze_r + evaluate_r) / cog_total if cog_total > 0 else 0.0

        # Difficulty levels
        l2_r = safe_int(row.get('Diff:L2 (Right)', ''))
        l2_t = safe_int(row.get('Diff:L2 (Total)', ''))
        l3_r = safe_int(row.get('Diff:L3 (Right)', ''))
        l3_t = safe_int(row.get('Diff:L3 (Total)', ''))

        l2_acc = l2_r / l2_t if l2_t > 0 else 0.0
        l3_acc = l3_r / l3_t if l3_t > 0 else 0.0
        cpd2 = (l2_acc + l3_acc) / 2.0

        # Behavioral metrics
        indecision = safe_float(row.get('Indecision Score', ''))
        flip_flops = safe_int(row.get('Flip-Flops', ''))
        max_indecision = max(max_indecision, indecision)

        user_records.append({
            'branchCode': bc,
            'cpd1': cpd1,
            'cpd2': cpd2,
            'indecision': indecision,
            'flip_flops': flip_flops,
            'total_q': cog_total,
        })

    if max_indecision == 0:
        max_indecision = 1.0

    # Second pass: compute CPD3, CPD4, CPI
    branch_cpis = defaultdict(list)
    for u in user_records:
        cpd3 = 1.0 - (u['indecision'] / max_indecision)
        cpd4 = 1.0 - (u['flip_flops'] / u['total_q']) if u['total_q'] > 0 else 1.0
        cpi = (u['cpd1'] + u['cpd2'] + cpd3 + cpd4) / 4.0
        if u['branchCode']:
            branch_cpis[u['branchCode']].append(cpi)

    scores = {}
    for bc, cpis in branch_cpis.items():
        cpi_avg = sum(cpis) / len(cpis)
        c3 = 20 * cpi_avg
        scores[bc] = {
            'C3_CPI':         round(cpi_avg, 4),
            'C3_Respondents': len(cpis),
            'C3':             round(c3, 2),
        }

    return scores


# ═══════════════════════════════════════════════════════════
#  C4 — System Readiness (max 20)
# ═══════════════════════════════════════════════════════════

def compute_c4():
    """
    Likert → binary:  SA/A (opt 0,1) → 1,  D/SD (opt 2,3) → 0
    EnablementIndex = mean(A1_mean, A2_mean, A3_mean, A4_mean)
    RoutineIndex    = mean(B1_mean, B2_mean, B3_mean, B4_mean, B5_mean)
    C4 = 10 × EnablementIndex  +  10 × RoutineIndex
    """
    # Build questionId → area code mapping from questions file
    q_area = {}
    q_path = os.path.join(STAT, 'system_readiness_questions.csv')
    for row in read_csv(q_path):
        qid  = row['questionId'].strip()
        desc = row.get('description', '').strip()
        if desc and len(desc) >= 2:
            q_area[qid] = desc[:2]  # "A1", "A2", ... "B5"

    # Accumulate per branch per area
    # branch_area[bc][area] = [binary_values]
    branch_area = defaultdict(lambda: defaultdict(list))
    r_path = os.path.join(STAT, 'system_readiness_responses_detail.csv')

    for row in read_csv(r_path):
        bc  = row['branchCode'].strip()
        qid = row['questionId'].strip()
        opt = safe_int(row.get('selectedOption', ''), -1)
        if opt < 0 or not bc:
            continue
        area = q_area.get(qid)
        if not area:
            continue
        binary = 1 if opt <= 1 else 0
        branch_area[bc][area].append(binary)

    scores = {}
    for bc in branch_area:
        a_means = {}
        b_means = {}
        for area, vals in branch_area[bc].items():
            m = sum(vals) / len(vals)
            if area.startswith('A'):
                a_means[area] = m
            elif area.startswith('B'):
                b_means[area] = m

        enablement = sum(a_means.values()) / len(a_means) if a_means else 0.0
        routine    = sum(b_means.values()) / len(b_means) if b_means else 0.0
        c4 = 10 * enablement + 10 * routine

        scores[bc] = {
            'C4_Enablement':  round(enablement, 4),
            'C4_Routine':     round(routine, 4),
            'C4':             round(c4, 2),
        }

    return scores


# ═══════════════════════════════════════════════════════════
#  Assembly
# ═══════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  SRI Branch-Level Score Computation")
    print("=" * 60)

    print("\nC1: Intent Readiness (max 20)")
    c1 = compute_c1()
    print(f"  Branches scored: {len(c1)}")

    print("\nC2: Practice Readiness (max 25)")
    c2 = compute_c2()
    print(f"  Branches scored: {len(c2)}")

    print("\nC3: Capacity Readiness (max 20)")
    c3 = compute_c3()
    print(f"  Branches scored: {len(c3)}")

    print("\nC4: System Readiness (max 20)")
    c4 = compute_c4()
    print(f"  Branches scored: {len(c4)}")

    print("\nC5: Ecosystem Readiness (max 15)")
    print("  No data — all branches get 0")

    # ── Outer join all constructs on branchCode ──
    all_bc = sorted(set(c1) | set(c2) | set(c3) | set(c4))
    code_to_name = build_code_to_name()

    C1_EMPTY = {'C1_TeacherScore': '', 'C1_LeaderScore': '', 'C1_Alignment': '', 'C1': ''}
    C2_EMPTY = {'C2_DepthEffectiveness': '', 'C2': ''}
    C3_EMPTY = {'C3_CPI': '', 'C3_Respondents': '', 'C3': ''}
    C4_EMPTY = {'C4_Enablement': '', 'C4_Routine': '', 'C4': ''}

    fields = [
        'BranchCode', 'BranchName', 'SchoolName',
        'C1_TeacherScore', 'C1_LeaderScore', 'C1_Alignment', 'C1',
        'C2_DepthEffectiveness', 'C2',
        'C3_CPI', 'C3_Respondents', 'C3',
        'C4_Enablement', 'C4_Routine', 'C4',
        'C5', 'SRI_Total', 'SRI_Max',
    ]

    csv_path = os.path.join(OUT, 'sri_branch_scores.csv')
    json_records = []
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for bc in all_bc:
            bn, sn = code_to_name.get(bc, ('', ''))
            row = {'BranchCode': bc, 'BranchName': bn, 'SchoolName': sn}
            row.update(c1.get(bc, C1_EMPTY))
            row.update(c2.get(bc, C2_EMPTY))
            row.update(c3.get(bc, C3_EMPTY))
            row.update(c4.get(bc, C4_EMPTY))
            row['C5'] = 0

            # SRI total (sum of available scores)
            vals = []
            for key in ('C1', 'C2', 'C3', 'C4'):
                v = row.get(key, '')
                vals.append(float(v) if v != '' else 0.0)
            row['SRI_Total'] = round(sum(vals), 2)
            row['SRI_Max'] = 85

            w.writerow(row)

            # Build JSON record (convert '' to None for JSON)
            jrow = {}
            for k, v in row.items():
                if v == '':
                    jrow[k] = None
                elif isinstance(v, str):
                    try:
                        jrow[k] = float(v)
                        if jrow[k] == int(jrow[k]):
                            jrow[k] = int(jrow[k])
                    except ValueError:
                        jrow[k] = v
                else:
                    jrow[k] = v
            json_records.append(jrow)

    # ── Write JSON data ──
    json_path = os.path.join(OUT, 'sri_dashboard_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({"main": json_records}, f, ensure_ascii=False)
    print(f"\n{'─' * 60}")
    print(f"CSV written: {csv_path}")
    print(f"JSON written: {json_path}")
    print(f"Total branches: {len(all_bc)}")

    # ── Quick summary ──
    print(f"\n{'─' * 60}")
    print("Score ranges:")
    for label, scores, mx in [('C1', c1, 20), ('C2', c2, 25), ('C3', c3, 20), ('C4', c4, 20)]:
        if scores:
            vals = [s[label] for s in scores.values() if s[label] != '']
            lo = min(vals)
            hi = max(vals)
            avg = sum(vals) / len(vals)
            print(f"  {label} (max {mx:2d}): {lo:6.2f} – {hi:6.2f}  (avg {avg:.2f}, n={len(vals)})")

    # ── Add to Excel workbook ──
    wb_path = os.path.join(OUT, 'SRI_All_Constructs_Data.xlsx')
    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Remove existing sheet if present
    if 'SRI Scores' in wb.sheetnames:
        del wb['SRI Scores']

    ws = wb.create_sheet(title='SRI Scores')

    hdr_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for ri, row in enumerate(reader, 1):
            for ci, val in enumerate(row, 1):
                # Try to parse as number
                try:
                    nval = float(val)
                    cell_val = int(nval) if nval == int(nval) else nval
                except (ValueError, TypeError):
                    cell_val = val
                cell = ws.cell(row=ri, column=ci, value=cell_val)
                cell.border = thin
                if ri == 1:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Auto-fit column widths
    for ci in range(1, ws.max_column + 1):
        mx_len = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
            mx_len = max(mx_len, len(str(row[0])) if row[0] else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(mx_len + 3, 55)

    ws.freeze_panes = 'A2'
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(wb_path)
    print(f"\nWorkbook updated: {wb_path}")
    print(f"Sheet 'SRI Scores' — {ws.max_row - 1} data rows")

    # ── Update SRI dashboard HTML with new data ──
    html_path = os.path.join(OUT, 'sri_dashboard.html')
    if os.path.exists(html_path):
        with open(html_path, 'r', encoding='utf-8') as f:
            html = f.read()

        # Find and replace the DATA array
        marker = 'const DATA='
        idx = html.index(marker)
        end_idx = html.index(';', idx)
        new_data = marker + json.dumps(json_records, ensure_ascii=False)
        html = html[:idx] + new_data + html[end_idx:]

        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"\nHTML dashboard updated: {html_path}")
    else:
        print(f"\nHTML dashboard not found at {html_path} — skipped")


if __name__ == '__main__':
    main()
