#!/usr/bin/env python3
"""Build a single Excel workbook with all SRI construct data as separate worksheets."""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), '..')
STAT = os.path.join(BASE, 'assets', 'myelin_stat_ro')
OUT = os.path.join(BASE, 'output')

# Sheet name max 31 chars in Excel
SHEETS = [
    # Construct 1 & 2: Intent / Practice — Teacher English
    ('C1 Tchr EN Qs',       os.path.join(STAT, 'intent_teacher_english_questions.csv')),
    ('C1 Tchr EN Attempts',  os.path.join(STAT, 'intent_teacher_english_attempts.csv')),
    ('C1 Tchr EN Responses', os.path.join(STAT, 'intent_teacher_english_responses.csv')),
    # Construct 1 & 2: Intent / Practice — Teacher Marathi
    ('C1 Tchr MR Qs',       os.path.join(STAT, 'intent_teacher_marathi_questions.csv')),
    ('C1 Tchr MR Attempts',  os.path.join(STAT, 'intent_teacher_marathi_attempts.csv')),
    ('C1 Tchr MR Responses', os.path.join(STAT, 'intent_teacher_marathi_responses.csv')),
    # Construct 1: Intent — Leader English
    ('C1 Ldr EN Qs',        os.path.join(STAT, 'intent_leader_english_questions.csv')),
    ('C1 Ldr EN Attempts',   os.path.join(STAT, 'intent_leader_english_attempts.csv')),
    ('C1 Ldr EN Responses',  os.path.join(STAT, 'intent_leader_english_responses.csv')),
    # Construct 1: Intent — Leader Marathi
    ('C1 Ldr MR Qs',        os.path.join(STAT, 'intent_leader_marathi_questions.csv')),
    ('C1 Ldr MR Attempts',   os.path.join(STAT, 'intent_leader_marathi_attempts.csv')),
    ('C1 Ldr MR Responses',  os.path.join(STAT, 'intent_leader_marathi_responses.csv')),
    # Construct 3: Capacity — MathTangle
    ('C3 MathTangle Master', os.path.join(OUT, 'mathangle_master.csv')),
    ('C3 MathTangle PerQ',   os.path.join(OUT, 'mathangle_per_question.csv')),
    ('C3 MathTangle Branch', os.path.join(OUT, 'mathangle_branch_summary.csv')),
    ('C3 MathTangle Exit',   os.path.join(OUT, 'mathangle_exit_levels.csv')),
    # Construct 4: System Readiness
    ('C4 System Qs',         os.path.join(STAT, 'system_readiness_questions.csv')),
    ('C4 System Attempts',   os.path.join(STAT, 'system_readiness_attempts.csv')),
    ('C4 System Responses',  os.path.join(STAT, 'system_readiness_responses_detail.csv')),
]

# Construct colors for header styling
CONSTRUCT_COLORS = {
    'C1': '1F4E79',  # Dark blue — Intent
    'C3': '2E7D32',  # Dark green — Capacity
    'C4': '7B1FA2',  # Dark purple — System
}

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

for sheet_name, csv_path in SHEETS:
    if not os.path.exists(csv_path):
        print(f"  SKIP (not found): {csv_path}")
        continue

    ws = wb.create_sheet(title=sheet_name)

    # Determine construct color
    prefix = sheet_name[:2]
    fill_color = CONSTRUCT_COLORS.get(prefix, '424242')
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Auto-fit column widths (cap at 50)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    row_count = ws.max_row - 1  # minus header
    print(f"  {sheet_name}: {row_count} rows")

# Add a TOC / Index sheet at the beginning
toc = wb.create_sheet(title='INDEX', index=0)
toc_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
toc_headers = ['Sheet', 'Construct', 'Description', 'Rows']
for col_idx, h in enumerate(toc_headers, 1):
    cell = toc.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = toc_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

toc_data = [
    ('C1 Tchr EN Qs',       '1-Intent',  'Teacher English — 22 questions'),
    ('C1 Tchr EN Attempts',  '1-Intent',  'Teacher English — 124 attempts (flat)'),
    ('C1 Tchr EN Responses', '1-Intent',  'Teacher English — response detail'),
    ('C1 Tchr MR Qs',       '1-Intent',  'Teacher Marathi — 22 questions'),
    ('C1 Tchr MR Attempts',  '1-Intent',  'Teacher Marathi — 504 attempts (flat)'),
    ('C1 Tchr MR Responses', '1-Intent',  'Teacher Marathi — response detail'),
    ('C1 Ldr EN Qs',        '1-Intent',  'Leader English — 20 questions'),
    ('C1 Ldr EN Attempts',   '1-Intent',  'Leader English — 20 attempts (flat)'),
    ('C1 Ldr EN Responses',  '1-Intent',  'Leader English — response detail'),
    ('C1 Ldr MR Qs',        '1-Intent',  'Leader Marathi — 20 questions'),
    ('C1 Ldr MR Attempts',   '1-Intent',  'Leader Marathi — 69 attempts (flat)'),
    ('C1 Ldr MR Responses',  '1-Intent',  'Leader Marathi — response detail'),
    ('C3 MathTangle Master', '3-Capacity','109 users with scores & cognitive breakdown'),
    ('C3 MathTangle PerQ',   '3-Capacity','Per-question responses with subtopic & level'),
    ('C3 MathTangle Branch', '3-Capacity','Branch-level summary (29 branches)'),
    ('C3 MathTangle Exit',   '3-Capacity','Exit level profiles (109 users)'),
    ('C4 System Qs',         '4-System',  '68 baseline questions (A1-A4, B1-B5)'),
    ('C4 System Attempts',   '4-System',  '676 attempts (flat)'),
    ('C4 System Responses',  '4-System',  'Response detail with option text'),
]

for r_idx, (sname, construct, desc) in enumerate(toc_data, 2):
    # Get row count from the actual sheet
    ws_ref = wb[sname] if sname in wb.sheetnames else None
    row_count = ws_ref.max_row - 1 if ws_ref else 0
    for col_idx, val in enumerate([sname, construct, desc, row_count], 1):
        cell = toc.cell(row=r_idx, column=col_idx, value=val)
        cell.border = thin_border
        # Color-code by construct
        if construct.startswith('1'):
            cell.fill = PatternFill(start_color='D6E4F0', fill_type='solid')
        elif construct.startswith('3'):
            cell.fill = PatternFill(start_color='D5E8D4', fill_type='solid')
        elif construct.startswith('4'):
            cell.fill = PatternFill(start_color='E1D5E7', fill_type='solid')

toc.column_dimensions['A'].width = 25
toc.column_dimensions['B'].width = 14
toc.column_dimensions['C'].width = 50
toc.column_dimensions['D'].width = 10
toc.freeze_panes = 'A2'

out_path = os.path.join(OUT, 'SRI_All_Constructs_Data.xlsx')
wb.save(out_path)
print(f"\nWorkbook saved: {out_path}")
print(f"Total sheets: {len(wb.sheetnames)}")
